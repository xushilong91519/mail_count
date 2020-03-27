from win32com.client import Dispatch
from datetime import datetime,timedelta
from copy import deepcopy
from openpyxl import Workbook
import shutil
import os

global PRODUCT
global NEW_DIR
global PLIST

PRODUCT={}
NEW_DIR="C:\Monthly_Report\\"
PLIST=[]

Local_Prod_List=os.path.join(os.path.curdir,'Product_List.txt')

print 'Syncing Product list from remote server...'

try:
    shutil.copy('\\\\nascnsh102v1\\CNSH2_MLSKoreanSupport\\Email_Tool\\Joker_Report\\Product_List.txt',Local_Prod_List)
    print 'Syncing successfully.'

except:
    print 'Syncing Failed, using local copy...'

with open(Local_Prod_List,'r') as in_file:
    item=in_file.readline().rstrip('\n')
    while item:
        PLIST.append(item)
        PRODUCT[item]=[0,0]
        item=in_file.readline().rstrip('\n')


########################################################
#          Function to get next email item.            #
########################################################

def Get_Next_EmailAndTime(emails):
    email=emails.GetPrevious()
    while not hasattr(email,'ReceivedTime'):
        email=emails.GetPrevious()
    rtime=datetime.strptime(str(email.ReceivedTime),'%m/%d/%y %H:%M:%S')
    return email,rtime
    
########################################################
#         Function to grab email by category           #
########################################################

def Grab_Email(mailbox,start_date,end_date):
    inbox=mailbox.Folders.Item("Inbox")
    emails=inbox.Items
    email=emails.GetLast()
    rtime=datetime.strptime(str(email.ReceivedTime),'%m/%d/%y %H:%M:%S')
    
    #point to the correct date    
    while rtime>end_date:
        email,rtime=Get_Next_EmailAndTime(emails)

    #start the email query
    report={}
    months=[]
    while rtime>start_date:
        print rtime
        if rtime.month<10:
            month='%s_0%s'%(rtime.year,rtime.month)
        else:
            month='%s_%s'%(rtime.year,rtime.month)
        product=''
        new=0
        isnoaction=False
        for cate in email.Categories.encode('utf8').split(', '):
            if cate=='No Action Needed':
                isnoaction=True
            elif cate in PLIST:
                product=cate
            elif cate=='New SR':
                new=1
            else:
                continue
        if isnoaction:
            email,rtime=Get_Next_EmailAndTime(emails)
            continue
        if product=='':
            product='Other Product'
        if not month in months:
            report[month]=deepcopy(PRODUCT)
            months.append(month)
        report[month][product][0]+=1
        report[month][product][1]+=new
        email,rtime=Get_Next_EmailAndTime(emails)
    
    return report,months


########################################################
#      Function to write export report to a Excel      #
########################################################

def ReportToExcel(report,months):
    months.sort()
    wb=Workbook()
    ws=wb.active
    ws.title="Total"
    ws.column_dimensions["A"].width=15
    ws['A1']="Month"
    
    LINE=ord('B')
    for month in months:
        ws['%s1'%chr(LINE)]=month
        LINE+=1
        
    ROW=2
    for product in PLIST:
        ws['A%s'%ROW]=product
        LINE=ord('B')
        for month in months:
            ws['%s%s'%(chr(LINE),ROW)]=report[month][product][0]
            LINE+=1
        ROW+=1
    ws['A%s'%ROW]="Total"
    LINE=ord('B')
    for month in months:
        ws['%s%s'%(chr(LINE),ROW)]=sum((x[0] for x in report[month].values()))
        LINE+=1

    ws=wb.create_sheet(title='New SR')
    ws.column_dimensions["A"].width=15
    ws['A1']="Month"
    
    LINE=ord('B')
    for month in months:
        ws['%s1'%chr(LINE)]=month
        LINE+=1
        
    ROW=2
    for product in PLIST:
        ws['A%s'%ROW]=product
        LINE=ord('B')
        for month in months:
            ws['%s%s'%(chr(LINE),ROW)]=report[month][product][1]
            LINE+=1
        ROW+=1
        
    ws['A%s'%ROW]="Total"
    LINE=ord('B')
    for month in months:
        ws['%s%s'%(chr(LINE),ROW)]=sum((x[1] for x in report[month].values()))
        LINE+=1
    ROW+=1

    try:
        wb.save(os.path.join(NEW_DIR,"Monthly_Report.xlsx"))
        print "A report was generated, please check %s"%(os.path.join(NEW_DIR,"Monthly_Report.xlsx"))
    except IOError:
        print "File Opened by some other process, Please run again after the current process finish...\n File:%s opened."%os.path.join(NEW_DIR,"Monthly_Report.xlsx")
    return 0



########################################################
#                   The Main() Fuction                 #
########################################################

def main():
    if not os.path.exists(NEW_DIR):
        outfile_path=os.mkdir(NEW_DIR)
    outlook=Dispatch("Outlook.Application").GetNamespace("MAPI")
    mailbox=''
    for account in outlook.Folders:
        account_name=account.Name.encode('utf8')
        if account_name.startswith("CSC_Korea"):
            mailbox=account_name
            break
        elif account_name.startswith("CSC Korea"):
            mailbox=account_name
    mailbox=outlook.Folders.Item(mailbox)
    print "This is the Monthly Report Generator for CSC Korea mailbox."
    start_date=raw_input("Please input the start date to query(Ex yyyy.mm.dd):")
    end_date=raw_input("Please input the end date to query(Ex yyyy.mm.dd),EMPTY for today:")
    print "Joker is grabing now, please wait..."
    start_date=datetime.strptime(start_date+' 00:00:01','%Y.%m.%d %H:%M:%S')
    if not end_date:
        end_date=datetime.now()
    else:
        end_date=datetime.strptime(end_date+' 23:59:59','%Y.%m.%d %H:%M:%S')

    Report,Months=Grab_Email(mailbox,start_date,end_date)
    
    ReportToExcel(Report,Months)

    END=raw_input("Press any key to exit...")

    del END

main()
