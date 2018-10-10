from win32com.client import Dispatch
from datetime import datetime,timedelta
from copy import deepcopy
from openpyxl import Workbook
import shutil
import os

global PRODUCT
global NEW_DIR
global PLIST

PRODUCT={}
NEW_DIR="C:\Mail_Report\\"
PLIST=[]

Local_Prod_List=os.path.join(os.path.curdir,'Product_List.txt')

print 'Syncing Product list from remote server...'

try:
    shutil.copy('\\\\nascnsh102v1\\CNSH2_MLSKoreanSupport\\Email_Tool\\Joker_Report\\Product_List.txt',Local_Prod_List)
    print 'Syncing successfully.'

except:
    print 'Syncing Failed, using local copy...'

with open(Local_Prod_List,'r') as in_file:
    item=in_file.readline().rstrip('\n')
    while item:
        PLIST.append(item)
        PRODUCT[item]=[0,0,0]
        item=in_file.readline().rstrip('\n')


########################################################
#          Function to get next email item.            #
########################################################

def Get_Next_EmailAndTime(emails):
    email=emails.GetPrevious()
    while not hasattr(email,'ReceivedTime'):
        email=emails.GetPrevious()
    rtime=datetime.strptime(str(email.ReceivedTime),'%m/%d/%y %H:%M:%S')
    return email,rtime
    
########################################################
#         Function to grab email by category           #
########################################################

def Grab_Email(mailbox,date):
    if date.day>datetime.now().day:
        print "Invalid date!"
        return 0
    inbox=mailbox.Folders.Item("Inbox")
    emails=inbox.Items
    email=emails.GetLast()
    end_time=date
    start_time=date-timedelta(days=1)
    rtime=datetime.strptime(str(email.ReceivedTime),'%m/%d/%y %H:%M:%S')
    
    #point to the correct date    
    while rtime>end_time:
        email,rtime=Get_Next_EmailAndTime(emails)

    #start the email query
    report={}
    while rtime>start_time:
        isnoaction=False
        agent=''
        product=''
        new=0
        se_var=0
        for cate in email.Categories.encode('utf8').split(', '):
            if cate=='No Action Needed':
                isnoaction=True
            elif cate in PRODUCT:
                product=cate
            elif cate=='New SR':
                new=1
            elif cate=='End User Direct Support' or cate=='VCE':
                continue
            elif cate=='SE-VAR':
                se_var=1
            else:
                agent=cate
        if isnoaction:
            email,rtime=Get_Next_EmailAndTime(emails)
            continue
        if agent=='':
            email,rtime=Get_Next_EmailAndTime(emails)
            continue
        if product=='':
            product='Other Product'
        if not agent in report:
            report[agent]=deepcopy(PRODUCT)
        report[agent][product][0]+=1
        report[agent][product][1]+=new
        report[agent][product][2]+=se_var
        email,rtime=Get_Next_EmailAndTime(emails)
            
    return report


########################################################
#      Function to write export report to a Excel      #
########################################################

def ReportToExcel(report,date):
    wb=Workbook()
    ws=wb.active
    ws.title="Total"
    ws.column_dimensions["A"].width=15
    ws['A1']="Product"
    
    LINE=ord('B')
    for agent in report.keys():
        ws['%s1'%chr(LINE)]=agent
        LINE+=1
        
    ROW=2
    for product in PLIST:
        ws['A%s'%ROW]=product
        LINE=ord('B')
        for agent in report.keys():
            ws['%s%s'%(chr(LINE),ROW)]=report[agent][product][0]
            LINE+=1
        ROW+=1
    ws['A%s'%ROW]="Total"
    LINE=ord('B')
    for agent in report:
        ws['%s%s'%(chr(LINE),ROW)]=sum((x[0] for x in report[agent].values()))
        LINE+=1

    ws=wb.create_sheet(title='New SR')
    ws.column_dimensions["A"].width=15
    ws['A1']="Product"
    
    LINE=ord('B')
    for agent in report.keys():
        ws['%s1'%chr(LINE)]=agent
        LINE+=1
        
    ROW=2
    for product in PLIST:
        ws['A%s'%ROW]=product
        LINE=ord('B')
        for agent in report.keys():
            if report[agent][product][2]:
                ws['%s%s'%(chr(LINE),ROW)]=report[agent][product][1]
            else:
                ws['%s%s'%(chr(LINE),ROW)]=report[agent][product][1]
            LINE+=1
        ROW+=1
        
    ws['A%s'%ROW]="Total"
    LINE=ord('B')
    for agent in report:
        ws['%s%s'%(chr(LINE),ROW)]=sum((x[1] for x in report[agent].values()))
        LINE+=1
    ROW+=1
    
    ws['A%s'%ROW]="SE-VAR"
    LINE=ord('B')
    for agent in report:
        SUM=sum((x[2] for x in report[agent].values()))
        sevar=""
        for product in report[agent]:
            if report[agent][product][2]:
                sevar=sevar+" %s:"%product+str(report[agent][product][2])
        ws['%s%s'%(chr(LINE),ROW)]="%s%s"%(SUM,sevar)
        LINE+=1

    try:
        wb.save(os.path.join(NEW_DIR,"Report_%s.xlsx"%date))
        print "A report was generated, please check %s"%(os.path.join(NEW_DIR,"Report_%s.xlsx"%date))
    except IOError:
        print "File Opened by some other process, Please run again after the current process finish...\n File:%s opened."%os.path.join(NEW_DIR,"Report_%s.xlsx"%date)

    return 0



########################################################
#                   The Main() Fuction                 #
########################################################

def main():
    if not os.path.exists(NEW_DIR):
        outfile_path=os.mkdir(NEW_DIR)
    outlook=Dispatch("Outlook.Application").GetNamespace("MAPI")
    mailbox=''
    for account in outlook.Folders:
        account_name=account.Name.encode('utf8')
        if account_name.startswith("CSC_Korea"):
            mailbox=account_name
            break
        elif account_name.startswith("CSC Korea"):
            mailbox=account_name
    mailbox=outlook.Folders.Item(mailbox)
    print "This is the Report Generator for CSC Korea mailbox."
    date=raw_input("Please input the date to query(Ex yyyy.mm.dd),Enter for today:")
    print "Joker is grabing now, please wait..."
    if not date:
        date=datetime.now().replace(hour=21,minute=0,second=0)
    else:
        date=datetime.strptime(date+" 21:00:00",'%Y.%m.%d %H:%M:%S')

    Report=Grab_Email(mailbox,date)
    
    ReportToExcel(Report,"%s_%s_%s"%(date.year,date.month,date.day))

    END=raw_input("Press any key to exit...")

    del END

main()
